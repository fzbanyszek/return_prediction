from __future__ import annotations

import hashlib
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

try:
    import psycopg
    from psycopg.types.json import Jsonb
except ImportError as exc:
    raise SystemExit(
        "Missing package psycopg"
    ) from exc

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing package openpyxl"
    ) from exc


DATA_ROOT = Path("raw_data")
DB_HOST = "localhost"
DB_PORT = 5432
DB_NAME = "stock_data"
DB_USER = "admin"
DB_PASSWORD = "admin"
BATCH_NAME = None
IF_EXISTS = "replace"
LIMIT_FILES = None
BATCH_SIZE = 500


HEADER_NORMALIZER = re.compile(r"[^0-9a-zA-Z]+")
GENERIC_FINANCIAL_TABLES = {
    "raw.income_statement",
    "raw.balance_sheet",
    "raw.cash_flow",
    "raw.enterprise_value",
    "raw.multiples",
    "raw.per_share",
    "raw.ratios_credit",
    "raw.ratios_liquidity",
    "raw.ratios_profitability",
    "raw.ratios_working_capital",
    "raw.ratios_yield_analysis",
}

PROFILE_FIELDS: list[tuple[str, str]] = [
    ("ticker", "ticker"),
    ("company_name", "company_name"),
    ("cik", "cik"),
    ("cusip", "cusip"),
    ("isin", "isin"),
    ("currency", "currency"),
    ("description", "description"),
    ("ai_description", "ai_description"),
    ("website", "website"),
    ("address", "address"),
    ("city", "city"),
    ("state", "state"),
    ("zip", "zip"),
    ("country", "country"),
    ("country_code", "country_code"),
    ("phone", "phone"),
    ("sector", "sector"),
    ("industry", "industry"),
    ("ceo", "ceo"),
    ("full_time_employees", "full_time_employees"),
    ("ipo_date", "ipo_date"),
    ("is_adr", "is_adr"),
    ("price", "price"),
    ("dividend_yield", "dividend_yield"),
    ("last_dividend", "last_dividend"),
    ("dividend_date", "dividend_date"),
    ("ex_dividend_date", "ex_dividend_date"),
    ("earnings_date", "earnings_date"),
    ("percentage_held_by_insiders", "percentage_held_by_insiders"),
    ("percentage_held_by_institutions", "percentage_held_by_institutions"),
    ("short_shares_outstanding", "short_shares_outstanding"),
    ("short_shares_outstanding_percentage", "short_shares_outstanding_percentage"),
    ("exchange_name", "exchange"),
    ("exchange_short_name", "exchange_short_name"),
]

PROFILE_INSERT_SQL = """
    insert into raw.company_profile (
        source_file_id,
        row_no,
        exchange_code,
        ticker,
        company_name,
        cik,
        cusip,
        isin,
        currency,
        description,
        ai_description,
        website,
        address,
        city,
        state,
        zip,
        country,
        country_code,
        phone,
        sector,
        industry,
        ceo,
        full_time_employees,
        ipo_date,
        is_adr,
        price,
        dividend_yield,
        last_dividend,
        dividend_date,
        ex_dividend_date,
        earnings_date,
        percentage_held_by_insiders,
        percentage_held_by_institutions,
        short_shares_outstanding,
        short_shares_outstanding_percentage,
        exchange_name,
        exchange_short_name,
        payload
    )
    values (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s
    )
"""

PRICE_INSERT_SQL = """
    insert into raw.price_daily (
        source_file_id,
        row_no,
        exchange_code,
        ticker,
        trade_date,
        open,
        high,
        low,
        close,
        adj_close,
        volume,
        unadjusted_volume,
        change,
        change_percent,
        vwap,
        label,
        payload
    )
    values (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
    )
"""

GENERIC_INSERT_SQL: dict[str, str] = {
    table_name: f"""
        insert into {table_name} (
            source_file_id,
            row_no,
            exchange_code,
            ticker,
            report_date,
            period,
            period_label,
            fiscal_year,
            currency,
            payload
        )
        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    for table_name in GENERIC_FINANCIAL_TABLES
}

TARGET_TABLE_BY_CATEGORY: dict[tuple[str, str | None], str] = {
    ("company_data", None): "raw.company_profile",
    ("price", None): "raw.price_daily",
    ("income_statement", None): "raw.income_statement",
    ("balance_sheet", None): "raw.balance_sheet",
    ("cash_flow", None): "raw.cash_flow",
    ("enterprise_value", None): "raw.enterprise_value",
    ("multiples", None): "raw.multiples",
    ("per_share", None): "raw.per_share",
    ("ratios", "credit"): "raw.ratios_credit",
    ("ratios", "liquidity"): "raw.ratios_liquidity",
    ("ratios", "profitability"): "raw.ratios_profitability",
    ("ratios", "working_capital"): "raw.ratios_working_capital",
    ("ratios", "yield_analysis"): "raw.ratios_yield_analysis",
}

REQUIRED_RAW_TABLES = (
    "import_batch",
    "source_file",
    "company_profile",
    "price_daily",
    "income_statement",
    "balance_sheet",
    "cash_flow",
    "enterprise_value",
    "multiples",
    "per_share",
    "ratios_credit",
    "ratios_liquidity",
    "ratios_profitability",
    "ratios_working_capital",
    "ratios_yield_analysis",
)


@dataclass(frozen=True)
class FileDescriptor:
    path: Path
    exchange_code: str
    ticker: str
    file_category: str
    file_subcategory: str | None
    period_type: str | None
    target_table: str


def main() -> int:
    data_root = DATA_ROOT.resolve()

    if not data_root.exists():
        print(f"Missing input directory: {data_root}")
        return 1

    files = discover_excel_files(data_root)
    if LIMIT_FILES is not None:
        files = files[:LIMIT_FILES]

    if not files:
        print("No XLSX files found.")
        return 0

    batch_name = BATCH_NAME or f"raw_import_{datetime.now(timezone.utc):%Y%m%d_%H%M%S}"
    total_rows = 0
    batch_id: int | None = None

    conn = psycopg.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
    )

    try:
        ensure_required_tables_exist(conn)
        batch_id = create_import_batch(conn, batch_name, data_root)

        for index, descriptor in enumerate(files, start=1):
            inserted_rows = process_file(
                conn=conn,
                descriptor=descriptor,
                import_batch_id=batch_id,
                if_exists=IF_EXISTS,
                batch_size=BATCH_SIZE,
                file_index=index,
                file_count=len(files),
            )
            total_rows += inserted_rows
            conn.commit()

        mark_import_batch(conn, batch_id, "completed", None)
        conn.commit()
        print(f"Completed. Files: {len(files)}, Rows: {total_rows}")
        return 0
    except Exception as exc:
        conn.rollback()
        if batch_id is not None:
            mark_import_batch(conn, batch_id, "failed", str(exc))
            conn.commit()
        print(f"Failed: {exc}")
        return 1
    finally:
        conn.close()


def ensure_required_tables_exist(conn: psycopg.Connection) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            select table_name
            from information_schema.tables
            where table_schema = 'raw'
            """
        )
        existing_tables = {row[0] for row in cur.fetchall()}

    missing = [table_name for table_name in REQUIRED_RAW_TABLES if table_name not in existing_tables]
    if missing:
        missing_list = ", ".join(missing)
        raise RuntimeError(
            "Missing tables in schema raw. "
            "Run create_raw_tables_script first. "
            f"Missing: {missing_list}"
        )


def discover_excel_files(data_root: Path) -> list[FileDescriptor]:
    descriptors: list[FileDescriptor] = []
    for path in sorted(data_root.rglob("*.xlsx")):
        descriptor = parse_file_descriptor(data_root, path)
        if descriptor is not None:
            descriptors.append(descriptor)
    return descriptors


def parse_file_descriptor(data_root: Path, path: Path) -> FileDescriptor | None:
    rel_parts = path.relative_to(data_root).parts
    if len(rel_parts) < 5:
        return None

    exchange_code = rel_parts[0].upper()
    if exchange_code not in {"NASDAQ", "NYSE"}:
        return None

    if rel_parts[1].lower() != "data":
        return None

    ticker = rel_parts[2].upper()
    file_category = rel_parts[3].lower()
    file_subcategory = rel_parts[4].lower() if file_category == "ratios" and len(rel_parts) >= 6 else None

    target_table = TARGET_TABLE_BY_CATEGORY.get((file_category, file_subcategory))
    if target_table is None:
        return None

    return FileDescriptor(
        path=path.resolve(),
        exchange_code=exchange_code,
        ticker=ticker,
        file_category=file_category,
        file_subcategory=file_subcategory,
        period_type=infer_period_type(path.name),
        target_table=target_table,
    )


def infer_period_type(file_name: str) -> str | None:
    lowered = file_name.lower()
    if "_annual_" in lowered:
        return "annual"
    if "_quarterly_" in lowered:
        return "quarterly"
    if "_ttm_" in lowered:
        return "ttm"
    return None


def create_import_batch(conn: psycopg.Connection, batch_name: str, data_root: Path) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            insert into raw.import_batch (batch_name, source_root_path, status)
            values (%s, %s, 'started')
            returning import_batch_id
            """,
            (batch_name, str(data_root)),
        )
        batch_id = cur.fetchone()[0]
    conn.commit()
    return batch_id


def mark_import_batch(
    conn: psycopg.Connection, batch_id: int, status: str, notes: str | None
) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            update raw.import_batch
            set finished_at = now(),
                status = %s,
                notes = %s
            where import_batch_id = %s
            """,
            (status, notes, batch_id),
        )


def process_file(
    conn: psycopg.Connection,
    descriptor: FileDescriptor,
    import_batch_id: int,
    if_exists: str,
    batch_size: int,
    file_index: int,
    file_count: int,
) -> int:
    existing_count = count_existing_source_files(conn, descriptor.path)
    if existing_count and if_exists == "skip":
        print(f"[{file_index}/{file_count}] Skipped: {descriptor.path.name}")
        return 0

    if existing_count and if_exists == "replace":
        delete_existing_source_files(conn, descriptor.path)

    workbook = load_workbook(descriptor.path, read_only=True, data_only=True)
    try:
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        headers = read_headers(worksheet)
        file_stats = descriptor.path.stat()
        source_file_id = insert_source_file(
            conn=conn,
            import_batch_id=import_batch_id,
            descriptor=descriptor,
            sheet_name=sheet_name,
            headers=headers,
            file_stats=file_stats,
        )

        row_count = insert_sheet_rows(
            conn=conn,
            descriptor=descriptor,
            source_file_id=source_file_id,
            worksheet=worksheet,
            headers=headers,
            batch_size=batch_size,
        )
        update_source_file_after_load(conn, source_file_id, row_count)
        print(f"[{file_index}/{file_count}] Imported: {descriptor.path.name}")
        return row_count
    finally:
        workbook.close()


def count_existing_source_files(conn: psycopg.Connection, path: Path) -> int:
    with conn.cursor() as cur:
        cur.execute(
            "select count(*) from raw.source_file where file_path = %s",
            (str(path),),
        )
        return int(cur.fetchone()[0])


def delete_existing_source_files(conn: psycopg.Connection, path: Path) -> None:
    with conn.cursor() as cur:
        cur.execute(
            "delete from raw.source_file where file_path = %s",
            (str(path),),
        )


def read_headers(worksheet: Any) -> list[str]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        raise ValueError(f"Sheet {worksheet.title} is empty.")

    normalized = [normalize_header(value) for value in first_row]
    while normalized and normalized[-1] == "":
        normalized.pop()

    if not normalized:
        raise ValueError(f"Sheet {worksheet.title} has no headers.")

    seen: set[str] = set()
    for header in normalized:
        if not header:
            raise ValueError(f"Sheet {worksheet.title} has an empty header.")
        if header in seen:
            raise ValueError(f"Sheet {worksheet.title} has a duplicated header: {header}")
        seen.add(header)

    return normalized


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    header = str(value).strip().lower()
    header = HEADER_NORMALIZER.sub("_", header).strip("_")
    return header


def insert_source_file(
    conn: psycopg.Connection,
    import_batch_id: int,
    descriptor: FileDescriptor,
    sheet_name: str,
    headers: list[str],
    file_stats: Any,
) -> int:
    checksum = calculate_sha256(descriptor.path)
    modified_at = datetime.fromtimestamp(file_stats.st_mtime, tz=timezone.utc)

    with conn.cursor() as cur:
        cur.execute(
            """
            insert into raw.source_file (
                import_batch_id,
                exchange_code,
                ticker,
                file_category,
                file_subcategory,
                period_type,
                file_name,
                file_path,
                sheet_name,
                header_columns,
                file_size_bytes,
                file_checksum,
                source_modified_at
            )
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            returning source_file_id
            """,
            (
                import_batch_id,
                descriptor.exchange_code,
                descriptor.ticker,
                descriptor.file_category,
                descriptor.file_subcategory,
                descriptor.period_type,
                descriptor.path.name,
                str(descriptor.path),
                sheet_name,
                Jsonb(headers),
                file_stats.st_size,
                checksum,
                modified_at,
            ),
        )
        return int(cur.fetchone()[0])


def update_source_file_after_load(conn: psycopg.Connection, source_file_id: int, row_count: int) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            update raw.source_file
            set row_count = %s,
                loaded_at = now()
            where source_file_id = %s
            """,
            (row_count, source_file_id),
        )


def insert_sheet_rows(
    conn: psycopg.Connection,
    descriptor: FileDescriptor,
    source_file_id: int,
    worksheet: Any,
    headers: list[str],
    batch_size: int,
) -> int:
    insert_sql = resolve_insert_sql(descriptor.target_table)
    row_count = 0
    batch: list[tuple[Any, ...]] = []

    with conn.cursor() as cur:
        for excel_row_no, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            values = list(row_values[: len(headers)])
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))

            if is_blank_row(values):
                continue

            row_dict = {header: value for header, value in zip(headers, values)}
            payload = build_payload(row_dict)
            insert_values = build_insert_values(
                descriptor=descriptor,
                source_file_id=source_file_id,
                row_no=excel_row_no,
                row_dict=row_dict,
                payload=payload,
            )
            batch.append(insert_values)

            if len(batch) >= batch_size:
                cur.executemany(insert_sql, batch)
                row_count += len(batch)
                batch.clear()

        if batch:
            cur.executemany(insert_sql, batch)
            row_count += len(batch)

    return row_count


def resolve_insert_sql(target_table: str) -> str:
    if target_table == "raw.company_profile":
        return PROFILE_INSERT_SQL
    if target_table == "raw.price_daily":
        return PRICE_INSERT_SQL
    try:
        return GENERIC_INSERT_SQL[target_table]
    except KeyError as exc:
        raise ValueError(f"Unsupported target_table: {target_table}") from exc


def build_insert_values(
    descriptor: FileDescriptor,
    source_file_id: int,
    row_no: int,
    row_dict: dict[str, Any],
    payload: dict[str, Any],
) -> tuple[Any, ...]:
    if descriptor.target_table == "raw.company_profile":
        return build_company_profile_values(
            descriptor=descriptor,
            source_file_id=source_file_id,
            row_no=row_no,
            row_dict=row_dict,
            payload=payload,
        )
    if descriptor.target_table == "raw.price_daily":
        return build_price_values(
            descriptor=descriptor,
            source_file_id=source_file_id,
            row_no=row_no,
            row_dict=row_dict,
            payload=payload,
        )
    return build_generic_financial_values(
        descriptor=descriptor,
        source_file_id=source_file_id,
        row_no=row_no,
        row_dict=row_dict,
        payload=payload,
    )


def build_company_profile_values(
    descriptor: FileDescriptor,
    source_file_id: int,
    row_no: int,
    row_dict: dict[str, Any],
    payload: dict[str, Any],
) -> tuple[Any, ...]:
    values: list[Any] = [source_file_id, row_no, descriptor.exchange_code]
    values.append(to_text(row_dict.get("ticker")) or descriptor.ticker)

    for dest_column, source_column in PROFILE_FIELDS[1:]:
        _ = dest_column
        values.append(to_text(row_dict.get(source_column)))

    values.append(Jsonb(payload))
    return tuple(values)


def build_price_values(
    descriptor: FileDescriptor,
    source_file_id: int,
    row_no: int,
    row_dict: dict[str, Any],
    payload: dict[str, Any],
) -> tuple[Any, ...]:
    return (
        source_file_id,
        row_no,
        descriptor.exchange_code,
        descriptor.ticker,
        to_date(row_dict.get("date")),
        to_decimal(row_dict.get("open")),
        to_decimal(row_dict.get("high")),
        to_decimal(row_dict.get("low")),
        to_decimal(row_dict.get("close")),
        to_decimal(row_dict.get("adj_close")),
        to_int(row_dict.get("volume")),
        to_int(row_dict.get("unadjusted_volume")),
        to_decimal(row_dict.get("change")),
        to_decimal(row_dict.get("change_percent")),
        to_decimal(row_dict.get("vwap")),
        to_text(row_dict.get("label")),
        Jsonb(payload),
    )


def build_generic_financial_values(
    descriptor: FileDescriptor,
    source_file_id: int,
    row_no: int,
    row_dict: dict[str, Any],
    payload: dict[str, Any],
) -> tuple[Any, ...]:
    return (
        source_file_id,
        row_no,
        descriptor.exchange_code,
        to_text(row_dict.get("ticker")) or descriptor.ticker,
        to_date(row_dict.get("date")),
        to_text(row_dict.get("period")) or descriptor.period_type,
        to_text(row_dict.get("period_label")),
        to_int(row_dict.get("fiscal_year")),
        to_text(row_dict.get("currency")),
        Jsonb(payload),
    )


def build_payload(row_dict: dict[str, Any]) -> dict[str, Any]:
    return {key: to_json_compatible(value) for key, value in row_dict.items()}


def to_json_compatible(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, (int, bool, str)):
        return value
    return str(value)


def to_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return Decimal(str(value))

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def to_int(value: Any) -> int | None:
    decimal_value = to_decimal(value)
    if decimal_value is None:
        return None
    return int(decimal_value)


def to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for parser in (
        lambda raw: datetime.fromisoformat(raw).date(),
        lambda raw: datetime.strptime(raw, "%Y-%m-%d").date(),
        lambda raw: datetime.strptime(raw, "%m/%d/%Y").date(),
    ):
        try:
            return parser(text)
        except ValueError:
            continue
    return None


def is_blank_row(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def calculate_sha256(path: Path) -> str:
    sha256 = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


if __name__ == "__main__":
    raise SystemExit(main())
